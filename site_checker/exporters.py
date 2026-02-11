import csv
import io
from typing import Iterable, List

from .checks import CSV_COLUMNS_BASE

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


def rows_to_csv_bytes(rows: Iterable[dict]) -> bytes:
    rows_list = list(rows)
    if not rows_list:
        fieldnames = CSV_COLUMNS_BASE
    else:
        # Собираем все ключи из всех рядов для получения динамических Alt-колонок
        fieldnames = CSV_COLUMNS_BASE.copy()
        all_keys = set()
        for row in rows_list:
            all_keys.update(row.keys())
        # Добавляем Alt-колонки в правильном порядке
        alt_keys = sorted(
            [k for k in all_keys if k.startswith("Alt-")],
            key=lambda x: int(x.split("-")[1]),
        )
        fieldnames.extend(alt_keys)

    buffer = io.StringIO()
    writer = csv.DictWriter(buffer, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    for row in rows_list:
        writer.writerow(row)
    return buffer.getvalue().encode("utf-8-sig")


def rows_to_xlsx_bytes(rows: Iterable[dict]) -> bytes:
    """Экспортирует данные в формат Excel (.xlsx)"""
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl не установлена. Установите: pip install openpyxl")

    rows_list = list(rows)
    if not rows_list:
        fieldnames = CSV_COLUMNS_BASE
    else:
        # Собираем все ключи из всех рядов для получения динамических Alt-колонок
        fieldnames = CSV_COLUMNS_BASE.copy()
        all_keys = set()
        for row in rows_list:
            all_keys.update(row.keys())
        # Добавляем Alt-колонки в правильном порядке
        alt_keys = sorted(
            [k for k in all_keys if k.startswith("Alt-")],
            key=lambda x: int(x.split("-")[1]),
        )
        fieldnames.extend(alt_keys)

    wb = Workbook()
    ws = wb.active
    ws.title = "SEO Check Results"

    # Стили для заголовка
    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Добавляем заголовки
    for col_idx, fieldname in enumerate(fieldnames, start=1):
        cell = ws.cell(row=1, column=col_idx, value=fieldname)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        ws.column_dimensions[cell.column_letter].width = 18

    # Добавляем данные
    for row_idx, row_data in enumerate(rows_list, start=2):
        for col_idx, fieldname in enumerate(fieldnames, start=1):
            value = row_data.get(fieldname, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(
                horizontal="left", vertical="top", wrap_text=True
            )

    # Закрепляем заголовок
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def rows_to_headings_xlsx_bytes(
    rows: Iterable[dict],
    separator: str = " => ",
    enabled_headings: List[str] | None = None,
) -> bytes:
    """Экспортирует только выбранные заголовки H1-H6 в группировке по доменам.

    Args:
        rows: Итерируемые данные результатов
        separator: Разделитель между несколькими заголовками в одной ячейке
        enabled_headings: Список выбранных заголовков (например, ['H2', 'H3']).
                         Если None, используются все.
    """
    if not HAS_OPENPYXL:
        raise ImportError("openpyxl не установлена. Установите: pip install openpyxl")

    rows_list = list(rows)
    all_heading_keys = ["H1", "H2", "H3", "H4", "H5", "H6"]

    # Если enabled_headings не указан, используем все доступные
    if enabled_headings is None:
        heading_keys = all_heading_keys
    else:
        # Фильтруем и сохраняем порядок
        heading_keys = [h for h in all_heading_keys if h in enabled_headings]

    wb = Workbook()
    ws = wb.active
    ws.title = "Headings"

    header_fill = PatternFill(
        start_color="3B6EDC", end_color="3B6EDC", fill_type="solid"
    )
    subheader_fill = PatternFill(
        start_color="DCE6F8", end_color="DCE6F8", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF")
    subheader_font = Font(bold=True, color="1F2A44")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    if not rows_list:
        for col_idx, key in enumerate(heading_keys, start=1):
            cell = ws.cell(row=1, column=col_idx, value=key)
            cell.fill = subheader_fill
            cell.font = subheader_font
            cell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col_idx)].width = 22
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    domains_data = []
    max_rows = 0
    for row in rows_list:
        domain = row.get("URL", "")
        headings = {}
        for key in heading_keys:
            raw_value = row.get(key, "")
            if raw_value:
                parts = [p.strip() for p in raw_value.split(separator)]
                parts = [p for p in parts if p]
            else:
                parts = []
            headings[key] = parts
        domain_max = max((len(values) for values in headings.values()), default=0)
        max_rows = max(max_rows, domain_max)
        domains_data.append((domain, headings))

    # Заголовки доменов (строка 1) и H2/H3/... (строка 2)
    for idx, (domain, headings) in enumerate(domains_data):
        start_col = idx * len(heading_keys) + 1
        end_col = start_col + len(heading_keys) - 1

        ws.merge_cells(
            start_row=1, start_column=start_col, end_row=1, end_column=end_col
        )
        cell = ws.cell(row=1, column=start_col, value=domain)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

        for offset, key in enumerate(heading_keys):
            col = start_col + offset
            subcell = ws.cell(row=2, column=col, value=key.lower())
            subcell.fill = subheader_fill
            subcell.font = subheader_font
            subcell.alignment = header_alignment
            ws.column_dimensions[get_column_letter(col)].width = 22

    # Данные
    for row_offset in range(max_rows):
        excel_row = 3 + row_offset
        for idx, (_, headings) in enumerate(domains_data):
            start_col = idx * len(heading_keys) + 1
            for offset, key in enumerate(heading_keys):
                col = start_col + offset
                values = headings.get(key, [])
                value = values[row_offset] if row_offset < len(values) else ""
                cell = ws.cell(row=excel_row, column=col, value=value)
                cell.alignment = Alignment(
                    horizontal="left", vertical="top", wrap_text=True
                )

    ws.freeze_panes = "A3"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
